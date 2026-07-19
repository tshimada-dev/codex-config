import re
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook


SAMPLE = Path(__file__).with_name("estimator-synthetic-sample.xlsx")
ROOT_README = Path(__file__).parents[2] / "README.md"
SAMPLE_README = Path(__file__).with_name("README.md")
EXPECTED_SHEETS = [
    "00_結論",
    "01_内訳",
    "03_WBS",
    "10_AI補正",
    "18_親統合",
    "15_前提リスク",
]
SENSITIVE_PATTERNS = {
    "email": re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE),
    "aws_access_key": re.compile(r"\bAKIA[0-9A-Z]{16}\b"),
    "github_token": re.compile(r"\bgh[pousr]_[A-Za-z0-9]{20,}\b"),
    "openai_key": re.compile(r"\bsk-[A-Za-z0-9_-]{20,}\b"),
    "private_key": re.compile(r"-----BEGIN [A-Z ]*PRIVATE KEY-----"),
    "windows_path": re.compile(r"\b[A-Za-z]:\\[^\s]+"),
}


class EstimatorSyntheticSampleTest(unittest.TestCase):
    def setUp(self):
        self.assertTrue(SAMPLE.is_file(), f"missing sample: {SAMPLE}")

    def test_xlsx_package_and_expected_sheets(self):
        with zipfile.ZipFile(SAMPLE) as archive:
            self.assertIsNone(archive.testzip())
            self.assertIn("xl/workbook.xml", archive.namelist())

        workbook = load_workbook(SAMPLE, data_only=False, read_only=True)
        self.assertEqual(EXPECTED_SHEETS, workbook.sheetnames)
        self.assertIn("架空ケース", workbook["00_結論"]["A3"].value)
        self.assertIn("実在の組織・個人・金額を含みません", workbook["00_結論"]["A3"].value)

    def test_auditable_formulas_and_fixed_coefficients(self):
        workbook = load_workbook(SAMPLE, data_only=False, read_only=True)
        self.assertEqual("=SUM(D4:D9)", workbook["03_WBS"]["D10"].value)
        self.assertEqual("=SUM(E4:E9)", workbook["03_WBS"]["E10"].value)
        self.assertEqual("=ROUND(SUM(I4:I9),1)", workbook["10_AI補正"]["I10"].value)
        self.assertEqual(
            [0.70, 0.70, 0.85, 0.85, 0.95, 1.00],
            [workbook["10_AI補正"][f"G{row}"].value for row in range(4, 10)],
        )
        for row in range(4, 10):
            self.assertTrue(workbook["10_AI補正"][f"H{row}"].value.startswith("=ROUND("))
            self.assertIn("WBS根拠:", workbook["10_AI補正"][f"N{row}"].value)
            self.assertIn("係数根拠:", workbook["10_AI補正"][f"N{row}"].value)

    def test_content_is_synthetic_and_sensitive_pattern_free(self):
        workbook = load_workbook(SAMPLE, data_only=False, read_only=True)
        strings = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        strings.append(f"{sheet.title}!{cell.coordinate}: {cell.value}")
        with zipfile.ZipFile(SAMPLE) as archive:
            names = archive.namelist()
            self.assertNotIn("xl/vbaProject.bin", names)
            self.assertFalse(any(name.startswith("xl/externalLinks/") for name in names))
            package_text = "\n".join(
                archive.read(name).decode("utf-8", errors="ignore")
                for name in names
                if name.endswith((".xml", ".rels"))
            )
        corpus = "\n".join(strings) + "\n" + package_text

        self.assertIn("すべての名称・件数・工数は架空", corpus)
        self.assertIn("実案件の精度や一般的優位性を示さない", corpus)
        for name, pattern in SENSITIVE_PATTERNS.items():
            self.assertIsNone(pattern.search(corpus), f"sensitive pattern `{name}` found")

    def test_readme_links_resolve(self):
        root_text = ROOT_README.read_text(encoding="utf-8")
        sample_text = SAMPLE_README.read_text(encoding="utf-8")
        relative = "docs/samples/estimator-synthetic-sample.xlsx"
        self.assertIn(f"({relative})", root_text)
        self.assertIn("(estimator-synthetic-sample.xlsx)", sample_text)
        self.assertTrue((ROOT_README.parent / relative).is_file())


if __name__ == "__main__":
    unittest.main()
