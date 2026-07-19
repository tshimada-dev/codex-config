#!/usr/bin/env python3
"""Apply the codex-effort-estimator presentation workbook format.

The formatter is intentionally deterministic. It can take the raw estimator
workbook shape or an already formatted workbook, then rebuild the first two
sheets for readability, remove charts, renumber sheets by tab order, apply
stable widths/styles, and run lightweight validation.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter, range_boundaries
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required. Use the bundled Codex Python runtime or install openpyxl."
    ) from exc


COLORS = {
    "header": "1F4E79",
    "header_text": "FFFFFF",
    "subtle_header": "DDEBF7",
    "total": "E2F0D9",
    "assumption": "FFF2CC",
    "risk": "FCE4D6",
    "neutral": "F2F2F2",
    "border": "D9E2F3",
    "text": "1F1F1F",
    "muted": "666666",
}

PRESENTATION_LABELS = [
    "結論",
    "内訳",
    "規模根拠",
    "WBS",
    "PERT",
    "単価アンカー",
    "パラメトリック",
    "FP",
    "UCP",
    "トップダウン",
    "AI補正",
    "公共レビュー",
    "リスクモデル",
    "制約",
    "Discovery",
    "前提リスク",
    "類推補正",
    "Repo",
    "親統合",
]

LABEL_ALIASES = {
    "サマリー": "結論",
    "工程別": "内訳",
    "単価契約": "単価アンカー",
}

METHOD_NAMES = {
    "WBS",
    "単価アンカー",
    "パラメトリック",
    "FP",
    "UCP",
    "トップダウン",
    "トップダウン三点",
    "制約/容量",
    "制約",
    "リスクモデル",
    "コンポーネント単価",
}

AI_MULTIPLIERS = {
    "定型実装": (
        0.70,
        "WBS作成者が削減可能と判定。係数は参照定数で固定し、補正パス側で値引き裁量を持たない。",
    ),
    "コード隣接": (
        0.85,
        "実装補助は効くが設計判断・レビュー・結合確認が残るため中程度の固定係数を適用。",
    ),
    "複雑実装": (
        0.90,
        "複雑な業務ルールやデバッグはAI支援より検証・判断が支配的なため保守的に適用。",
    ),
    "検証重": (
        0.95,
        "帳票忠実度、旧新比較、受入証跡など検証中心の作業はほぼ削らない。",
    ),
    "削減不可": (
        1.00,
        "要件、受入、調整、説明責任などAIコーディングで削らない作業。",
    ),
    "対象外": (
        1.00,
        "AI補正対象外。raw baselineをそのまま保持。",
    ),
}

AI_TAG_ALIASES = {
    "削減あり": "コード隣接",
    "一部削減": "コード隣接",
    "削減困難": "削減不可",
    "削りすぎ注意": "検証重",
    "非削減": "削減不可",
}

WIDTHS = {
    "00_結論": [24, 32, 38, 16, 24, 24],
    "01_内訳": [24, 15, 15, 13, 12, 16, 58, 24],
    "02_規模根拠": [14, 28, 14, 54, 12, 38],
    "03_WBS": [16, 34, 54, 11, 13, 11, 14, 46],
    "04_PERT": [34, 48, 11, 13, 11, 12, 10, 11, 14, 42],
    "05_単価アンカー": [26, 9, 30, 11, 11, 11, 10, 10, 10, 18, 16, 11, 11, 11, 28, 54],
    "06_パラメトリック": [28, 12, 12, 12, 12, 58],
    "07_FP": [24, 12, 12, 12, 58],
    "08_UCP": [24, 12, 12, 12, 58],
    "09_トップダウン": [24, 12, 12, 12, 12, 64],
    "10_AI補正": [16, 34, 14, 11, 11, 11, 11, 12, 12, 12, 12, 18, 36, 54],
    "11_公共レビュー": [22, 58, 16, 54, 24],
    "12_リスクモデル": [28, 20, 62, 18],
    "13_制約": [24, 12, 22, 58],
    "14_Discovery": [28, 42, 11, 11, 11, 32, 46],
    "15_前提リスク": [12, 64, 42, 46],
    "16_類推補正": [26, 38, 38, 16, 13, 42],
    "17_Repo": [22, 54, 22, 11, 11, 11, 44],
    "18_親統合": [28, 10, 18, 18, 14, 18, 22, 64],
}

TAB_COLORS = {
    "00_結論": "70AD47",
    "01_内訳": "5B9BD5",
    "02_規模根拠": "9EADCC",
    "03_WBS": "4472C4",
    "04_PERT": "4472C4",
    "05_単価アンカー": "A9D18E",
    "06_パラメトリック": "A9D18E",
    "07_FP": "A9D18E",
    "08_UCP": "A9D18E",
    "09_トップダウン": "A9D18E",
    "10_AI補正": "F4B183",
    "11_公共レビュー": "F4B183",
    "12_リスクモデル": "C65911",
    "13_制約": "C65911",
    "14_Discovery": "FFD966",
    "15_前提リスク": "C65911",
    "16_類推補正": "B4C6E7",
    "17_Repo": "B4C6E7",
    "18_親統合": "B4C6E7",
}

NUMERIC_HEADERS = {
    "Low",
    "Most likely",
    "Likely",
    "High",
    "Base",
    "WBS",
    "PERT",
    "AI補正後",
    "親最終",
    "数量",
    "楽観",
    "最頻/普通",
    "悲観",
    "期待値",
    "SD",
    "分散",
    "ベースライン",
    "倍率",
    "補正後",
    "楽観/Low",
    "普通/Base",
    "悲観/High",
    "中心/期待値",
    "AI補助後目安",
    "AI補助前WBS",
    "差分",
    "削減率",
    "Raw Low",
    "Raw Base",
    "Raw High",
    "固定倍率",
    "Adjusted Low",
    "Adjusted Base",
    "Adjusted High",
    "Base差分",
}

FORMULA_ERRORS = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A"}

REPETITION_TERMS = (
    "地区",
    "帳票",
    "出力",
    "CSV",
    "変種",
    "類似",
    "バリアント",
    "variant",
    "report",
    "output",
)

REPETITION_EXCLUSION_TERMS = ("行", "ページ", "量", "容量", "データ行")

REPETITION_LABEL_EXCLUSIONS = ("資料ファイル", "外部システム", "対象環境")

REUSE_ASSUMPTION_TERMS = (
    "共通",
    "再利用",
    "パラメータ",
    "テンプレート",
    "skeleton",
    "variant",
    "バリアント",
    "reuse",
)

REUSE_FACTOR_PATTERNS = (
    re.compile(r"(?:variant|reuse|factor|係数|倍率)[^\d]*(0\.\d+(?:\s*[-～]\s*0\.\d+)?)", re.IGNORECASE),
    re.compile(r"(?:×|x)\s*(0\.\d+)", re.IGNORECASE),
)

AI_IMPLEMENTATION_TERMS = (
    "実装",
    "基盤",
    "UI",
    "画面",
    "CRUD",
    "CSV",
    "取込",
    "出力",
    "帳票",
    "PDF",
    "Excel",
    "テンプレ",
    "マスタ",
    "計算",
    "判定",
    "採番",
    "コード",
    "VBA",
    "script",
    "parser",
    "mapping",
)

AI_REVIEW_HEAVY_TAGS = {"複雑実装", "検証重"}

NON_ADDITIVE_HEADERS = {
    "",
    "倍率",
    "固定倍率",
    "削減率",
    "確率",
    "係数",
    "集中係数",
    "SD",
    "Total SD",
    "90% CI Low",
    "90% CI High",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Normalize a codex-effort-estimator .xlsx workbook format."
    )
    parser.add_argument("workbook", type=Path, help="Input .xlsx workbook.")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output path. Defaults to <input stem>_formatted.xlsx.",
    )
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="Overwrite the input workbook instead of writing a formatted copy.",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Exit non-zero when validation warnings are found.",
    )
    return parser.parse_args()


def fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def label_for(title: str) -> str:
    label = re.sub(r"^\d+_", "", title)
    return LABEL_ALIASES.get(label, label)


d…12968 tokens truncated…ct[str, str] = {}
    max_ai_row, _ = used_bounds(ai_ws)
    for row in range(ai_header + 1, max_ai_row + 1):
        task = text(ai_ws.cell(row, ai_headers["WBS作業"]).value)
        if not task or task == "合計":
            continue
        ai_tags[task] = normalize_ai_tag(ai_ws.cell(row, ai_headers["AI削減区分"]).value)

    max_breakdown_row, _ = used_bounds(breakdown_ws)
    for row in range(breakdown_header + 1, max_breakdown_row + 1):
        task = text(breakdown_ws.cell(row, breakdown_label_col).value)
        if not task or task == "合計" or task not in ai_tags:
            continue
        breakdown_tag = normalize_ai_tag(breakdown_ws.cell(row, breakdown_headers["AI削減区分"]).value)
        if breakdown_tag != ai_tags[task]:
            warnings.append(
                f"{breakdown_ws.title}!{get_column_letter(breakdown_headers['AI削減区分'])}{row}: "
                f"AI tag `{breakdown_tag}` does not match 10_AI補正 `{ai_tags[task]}` for `{task}`"
            )
    return warnings


def check_ai_reducibility_bias(wb: Any) -> list[str]:
    warnings: list[str] = []
    ws = sheet_by_label(wb, "AI補正")
    if ws is None:
        return warnings
    header_row, headers = find_header_row(
        ws,
        {"WBS作業", "AI削減区分", "Raw Base", "Adjusted Base"},
    )
    if header_row is None:
        return warnings

    total_raw = 0.0
    total_adjusted = 0.0
    implementation_raw = 0.0
    review_heavy_raw = 0.0
    review_heavy_examples: list[str] = []
    max_row, _ = used_bounds(ws)
    for row in range(header_row + 1, max_row + 1):
        task = text(ws.cell(row, headers["WBS作業"]).value)
        if not task or task == "合計":
            continue
        raw = numeric_cell_value(ws, row, headers["Raw Base"])
        adjusted = numeric_cell_value(ws, row, headers["Adjusted Base"])
        if raw is None or raw <= 0:
            continue
        if adjusted is not None:
            total_raw += raw
            total_adjusted += adjusted
        haystack = " ".join(
            [
                task,
                text(ws.cell(row, headers.get("WBS分類", 0)).value) if "WBS分類" in headers else "",
                text(ws.cell(row, headers.get("根拠", 0)).value) if "根拠" in headers else "",
            ]
        )
        if not contains_any(haystack, AI_IMPLEMENTATION_TERMS):
            continue
        implementation_raw += raw
        tag = normalize_ai_tag(ws.cell(row, headers["AI削減区分"]).value)
        if tag in AI_REVIEW_HEAVY_TAGS:
            review_heavy_raw += raw
            if len(review_heavy_examples) < 4:
                review_heavy_examples.append(f"{task}={tag}")

    if not total_raw or not implementation_raw:
        return warnings
    reduction_rate = (total_raw - total_adjusted) / total_raw if total_raw else 0.0
    review_heavy_share = review_heavy_raw / implementation_raw
    if review_heavy_share >= 0.70 and reduction_rate < 0.15:
        examples = ", ".join(review_heavy_examples)
        warnings.append(
            "AI reducibility sanity check: implementation-like work is mostly tagged "
            f"`複雑実装/検証重` ({review_heavy_share:.0%} of implementation raw base) while "
            f"overall AI reduction is only {reduction_rate:.0%}. Re-evaluate AI削減区分 against "
            f"the current scope and implementation approach. Examples: {examples}"
        )
    return warnings


def check_reuse_audit(wb: Any) -> list[str]:
    warnings: list[str] = []
    context = collect_reuse_context(wb)
    if not context["signals"]:
        return warnings

    component_ws = sheet_by_label(wb, "単価アンカー")
    if component_ws is None:
        warnings.append("repetition/reuse signals detected but 05_単価アンカー is missing")
    elif not component_anchor_is_detailed(component_ws):
        warnings.append(
            "repetition/reuse signals detected but 05_単価アンカー lacks count/framework/unit/variant-factor audit columns"
        )
    else:
        max_row, max_col = used_bounds(component_ws)
        header_row = likely_table_header_row(component_ws)
        headers = header_map(component_ws, header_row, max_col)
        factor_col = next(
            (
                col
                for header, col in headers.items()
                if "variant" in header.lower() or "reuse" in header.lower() or "factor" in header.lower() or "係数" in header
            ),
            None,
        )
        if factor_col:
            for row in range(header_row + 1, max_row + 1):
                family = text(component_ws.cell(row, 1).value)
                if not family or family in {"合計", "Total"}:
                    continue
                factor = text(component_ws.cell(row, factor_col).value)
                if factor.startswith("未記載"):
                    warnings.append(
                        f"{component_ws.title}!{get_column_letter(factor_col)}{row}: variant/reuse factor is missing for `{family}`"
                    )

    parent_ws = sheet_by_label(wb, "親統合")
    if parent_ws is None:
        warnings.append("repetition/reuse signals detected but 18_親統合 is missing")
    elif not parent_has_reuse_crosscheck(parent_ws):
        warnings.append(
            "repetition/reuse signals detected but 18_親統合 lacks the top-down per-unit reuse cross-check table"
        )
    else:
        max_row, max_col = used_bounds(parent_ws)
        for row in range(1, max_row + 1):
            headers = header_map(parent_ws, row, max_col)
            if "Variant/reuse factor" not in headers:
                continue
            factor_col = headers["Variant/reuse factor"]
            for check_row in range(row + 1, max_row + 1):
                if not text(parent_ws.cell(check_row, 1).value):
                    break
                factor = text(parent_ws.cell(check_row, factor_col).value)
                if factor.startswith("未記載"):
                    warnings.append(
                        f"{parent_ws.title}!{get_column_letter(factor_col)}{check_row}: variant/reuse factor is missing"
                    )
            break
    return warnings


def style_generic_sheet(ws: Any) -> None:
    st = base_styles()
    ws._charts = []
    ws.sheet_view.showGridLines = False
    max_row, max_col = used_bounds(ws)
    widths = WIDTHS.get(ws.title, [])
    for col in range(1, max(max_col, len(widths)) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1] if col <= len(widths) else 16
    if ws.title in {"00_結論", "01_内訳"}:
        for row_idx in range(1, max_row + 1):
            if ws.title == "00_結論":
                ws.row_dimensions[row_idx].height = 46 if 3 <= row_idx <= 8 else 30
            else:
                ws.row_dimensions[row_idx].height = 32
        return

    header_row = likely_table_header_row(ws)
    ws.freeze_panes = f"A{header_row + 1}"
    if max_row >= 1:
        for cell in ws[1]:
            cell.fill = fill(COLORS["header"])
            cell.font = st["section_font"]
            cell.alignment = st["left"]
            cell.border = st["header_border"]
        ws.row_dimensions[1].height = 28
    if max_row >= 2:
        for cell in ws[2]:
            cell.fill = fill(COLORS["neutral"])

    for row in range(1, max_row + 1):
        first_value = text(ws.cell(row, 1).value)
        row_is_header = row == header_row
        row_is_total = first_value in {"合計", "Total", "WBS由来合計"} or "最終推奨" in first_value
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if row == 1:
                cell.font = st["section_font"]
            elif row_is_header:
                cell.font = st["header_font"]
            elif row_is_total:
                cell.font = st["bold_font"]
            else:
                cell.font = st["body_font"]
            cell.border = st["border"]
            if row_is_header:
                cell.fill = fill(COLORS["header"])
                cell.font = Font(name="Yu Gothic", size=10, bold=True, color=COLORS["header_text"])
                cell.alignment = st["center"]
            elif row_is_total:
                cell.fill = fill(COLORS["total"])
                cell.alignment = st["right"] if isinstance(cell.value, (int, float)) else st["left"]
            else:
                cell.alignment = st["right"] if isinstance(cell.value, (int, float)) else st["left"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0"
        ws.row_dimensions[row].height = 28
    if max_col and max_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"


def normalize_presentation_workbook(wb: Any) -> None:
    reuse_context = collect_reuse_context(wb)
    source_summary = sheet_by_label(wb, "結論", "サマリー")
    source_breakdown = sheet_by_label(wb, "内訳", "工程別")
    source_wbs = sheet_by_label(wb, "WBS")
    source_ai = sheet_by_label(wb, "AI補正")
    source_component = sheet_by_label(wb, "単価アンカー")
    summary = extract_summary_values(source_summary)
    title = extract_summary_title(source_summary)
    methods = extract_method_rows(source_summary)
    ai_rows = extract_wbs_ai_rows(source_wbs)
    phases = phase_rows_from_ai_rows(ai_rows) if ai_rows else extract_phase_rows(source_breakdown)

    conclusion_ws = source_summary or wb.create_sheet("00_結論", 0)
    breakdown_ws = source_breakdown or wb.create_sheet("01_内訳", 1)
    ai_ws = source_ai or wb.create_sheet("10_AI補正")
    rebuild_conclusion_sheet(conclusion_ws, summary, methods, title)
    rebuild_breakdown_sheet(breakdown_ws, phases)
    if ai_rows:
        rebuild_ai_adjustment_sheet(ai_ws, ai_rows)
    if source_component is not None:
        enhance_component_anchor_sheet(source_component, reuse_context)
    renumber_and_order_sheets(wb)
    ensure_parent_reuse_crosscheck(wb, reuse_context)
    mark_wbs_derived_pert(wb)
    for ws in wb.worksheets:
        style_generic_sheet(ws)
        ws.sheet_properties.tabColor = TAB_COLORS.get(ws.title, COLORS["border"])


def set_full_recalc(wb: Any) -> None:
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True


def validate_workbook(wb: Any) -> tuple[list[str], list[str]]:
    warnings: list[str] = []
    errors: list[str] = []
    expected_present = [f"{idx:02d}_{label}" for idx, label in enumerate(PRESENTATION_LABELS)]
    seen_standard = [name for name in wb.sheetnames if name in expected_present]
    if seen_standard != [name for name in expected_present if name in seen_standard]:
        warnings.append("standard sheets are not in the expected presentation order")

    for required in ["00_結論", "01_内訳", "02_規模根拠", "03_WBS", "04_PERT", "18_親統合", "15_前提リスク"]:
        if required not in wb.sheetnames:
            warnings.append(f"missing expected sheet after formatting: {required}")

    for ws in wb.worksheets:
        if ws._charts:
            errors.append(f"{ws.title}: charts were not removed")
        errors.extend(check_total_crossfoot(ws))
        if ws.title == "10_AI補正":
            errors.extend(check_ai_adjustment_crossfoot(ws))
        if ws.title == "01_内訳":
            errors.extend(check_breakdown_crossfoot(ws))
        max_row, max_col = used_bounds(ws)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                value = text(ws.cell(row, col).value)
                if value in FORMULA_ERRORS:
                    errors.append(f"{ws.title}!{get_column_letter(col)}{row}: {value}")
    warnings.extend(check_reuse_audit(wb))
    warnings.extend(check_method_dependence_audit(wb))
    warnings.extend(check_breakdown_ai_tags(wb))
    warnings.extend(check_ai_reducibility_bias(wb))
    return warnings, errors


def main() -> int:
    args = parse_args()
    input_path = args.workbook
    if not input_path.exists():
        print(f"Workbook not found: {input_path}", file=sys.stderr)
        return 2
    if input_path.suffix.lower() != ".xlsx":
        print(f"Expected an .xlsx file: {input_path}", file=sys.stderr)
        return 2
    if args.output and args.in_place:
        print("Use either --output or --in-place, not both.", file=sys.stderr)
        return 2

    output_path = args.output
    if output_path is None:
        output_path = input_path if args.in_place else input_path.with_name(f"{input_path.stem}_formatted.xlsx")

    wb = load_workbook(input_path)
    set_full_recalc(wb)
    normalize_presentation_workbook(wb)
    warnings, errors = validate_workbook(wb)
    wb.active = 0
    wb.save(output_path)

    report = {
        "input": str(input_path),
        "output": str(output_path),
        "sheets": wb.sheetnames,
        "warnings": warnings,
        "errors": errors,
        "representative_visual_check_sheets": [
            name for name in ["00_結論", "01_内訳", "03_WBS", "10_AI補正", "18_親統合"] if name in wb.sheetnames
        ],
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))

    if errors or (args.strict and warnings):
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
