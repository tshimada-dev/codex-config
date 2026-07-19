import unittest

from openpyxl import Workbook

import format_estimate_workbook as formatter


AI_HEADERS = [
    "WBS分類",
    "WBS作業",
    "AI削減区分",
    "Raw Low",
    "Raw Base",
    "Raw High",
    "固定倍率",
    "Adjusted Low",
    "Adjusted Base",
    "Adjusted High",
    "Base差分",
    "判断者",
    "係数権限",
    "根拠",
]


class FormatEstimateWorkbookTest(unittest.TestCase):
    @staticmethod
    def workbook_with_ai_rows(rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "10_AI補正"
        sheet.append(AI_HEADERS)
        for row in rows:
            sheet.append(row)
        return workbook, sheet

    def test_fixed_line_level_coefficients_and_legacy_aliases(self):
        expected = {
            "定型実装": 0.70,
            "コード隣接": 0.85,
            "複雑実装": 0.90,
            "検証重": 0.95,
            "削減不可": 1.00,
            "対象外": 1.00,
        }

        self.assertEqual(expected, {tag: formatter.ai_multiplier_for(tag)[0] for tag in expected})
        self.assertEqual(0.85, formatter.ai_multiplier_for("削減あり")[0])
        self.assertEqual(1.00, formatter.ai_multiplier_for("未定義区分")[0])

    def test_crossfoot_preserves_raw_and_applies_fixed_multiplier(self):
        _, sheet = self.workbook_with_ai_rows(
            [
                [
                    "実装",
                    "CSV取込",
                    "定型実装",
                    8.0,
                    10.0,
                    12.0,
                    0.70,
                    5.6,
                    7.0,
                    8.4,
                    -3.0,
                    "WBS作成者",
                    "固定参照",
                    "薄いExcel/VBAスコープの定型処理",
                ]
            ]
        )

        self.assertEqual([], formatter.check_ai_adjustment_crossfoot(sheet))
        self.assertEqual(10.0, sheet["E2"].value)
        self.assertEqual(7.0, sheet["I2"].value)

    def test_crossfoot_rejects_discretionary_multiplier(self):
        _, sheet = self.workbook_with_ai_rows(
            [
                [
                    "実装",
                    "採番",
                    "定型実装",
                    8.0,
                    10.0,
                    12.0,
                    0.90,
                    7.2,
                    9.0,
                    10.8,
                    -1.0,
                    "WBS作成者",
                    "補正担当",
                    "目標値に合わせた任意係数",
                ]
            ]
        )

        errors = formatter.check_ai_adjustment_crossfoot(sheet)

        self.assertTrue(any("does not match fixed coefficient 0.70" in error for error in errors))

    def test_wbs_rebuild_preserves_scope_specific_line_rationale(self):
        workbook = Workbook()
        wbs_sheet = workbook.active
        wbs_sheet.title = "03_WBS"
        wbs_sheet.append(["分類", "作業", "Low", "Most likely", "High", "AI削減区分", "根拠"])
        wbs_sheet.append(
            [
                "実装",
                "CSV取込",
                8.0,
                10.0,
                12.0,
                "定型実装",
                "Excel/VBAで既存テンプレートへ取り込み、詳細テストは顧客が実施",
            ]
        )
        ai_sheet = workbook.create_sheet("10_AI補正")

        rows = formatter.extract_wbs_ai_rows(wbs_sheet)
        formatter.rebuild_ai_adjustment_sheet(ai_sheet, rows)

        self.assertEqual("Excel/VBAで既存テンプレートへ取り込み、詳細テストは顧客が実施", rows[0]["basis"])
        self.assertIn("WBS根拠: Excel/VBAで既存テンプレートへ取り込み、詳細テストは顧客が実施", ai_sheet["N5"].value)
        self.assertIn("係数根拠:", ai_sheet["N5"].value)

    def test_conservatism_warning_requires_scope_tag_recheck(self):
        workbook, _ = self.workbook_with_ai_rows(
            [
                ["実装", "CSV取込", "複雑実装", 80, 100, 120, 0.90, 72, 90, 108, -10, "WBS作成者", "固定参照", "VBA"],
                ["実装", "テンプレ埋め", "検証重", 40, 50, 60, 0.95, 38, 47.5, 57, -2.5, "WBS作成者", "固定参照", "既存テンプレ"],
                ["管理", "PM", "削減不可", 8, 10, 12, 1.00, 8, 10, 12, 0, "WBS作成者", "固定参照", "調整"],
            ]
        )

        warnings = formatter.check_ai_reducibility_bias(workbook)

        self.assertEqual(1, len(warnings))
        self.assertIn("Re-evaluate AI削減区分 against the current scope", warnings[0])
        self.assertIn("CSV取込=複雑実装", warnings[0])

    def test_routine_scope_tags_do_not_trigger_conservatism_warning(self):
        workbook, _ = self.workbook_with_ai_rows(
            [
                ["実装", "CSV取込", "定型実装", 80, 100, 120, 0.70, 56, 70, 84, -30, "WBS作成者", "固定参照", "VBA"],
                ["実装", "テンプレ埋め", "コード隣接", 40, 50, 60, 0.85, 34, 42.5, 51, -7.5, "WBS作成者", "固定参照", "既存テンプレ"],
                ["管理", "PM", "削減不可", 8, 10, 12, 1.00, 8, 10, 12, 0, "WBS作成者", "固定参照", "調整"],
            ]
        )

        self.assertEqual([], formatter.check_ai_reducibility_bias(workbook))

    @staticmethod
    def workbook_with_parent_methods():
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "18_親統合"
        sheet.append(["手法", "Low", "Base", "High"])
        sheet.append(["WBS", 1000, 1200, 1500])
        sheet.append(["UCP", 950, 1180, 1450])
        sheet.append(["FP", 650, 800, 950])
        sheet.append([])
        return workbook, sheet

    def test_empty_method_cluster_table_is_not_audit_evidence(self):
        workbook, sheet = self.workbook_with_parent_methods()
        sheet.append(
            [
                "Cluster",
                "Methods",
                "Shared assumptions",
                "Representative center",
                "Effective vote",
                "Independent anchors checked",
                "Anchor disposition",
                "Decision impact",
                "Reason",
            ]
        )

        warnings = formatter.check_method_dependence_audit(workbook)

        self.assertTrue(any("has no data rows" in warning for warning in warnings))

    def test_method_cluster_rows_require_one_vote_and_decision_effect(self):
        workbook, sheet = self.workbook_with_parent_methods()
        sheet.append(
            [
                "Cluster",
                "Methods",
                "Shared assumptions",
                "Representative center",
                "Effective vote",
                "Independent anchors checked",
                "Anchor disposition",
                "Decision impact",
                "Reason",
            ]
        )
        sheet.append(
            [
                "use-case/lifecycle",
                "WBS, UCP",
                "same use-case count and lifecycle",
                1190,
                "",
                "FP",
                "adopted",
                "",
                "best matches accepted delivery scope",
            ]
        )

        warnings = formatter.check_method_dependence_audit(workbook)

        self.assertTrue(any("Effective vote" in warning for warning in warnings))
        self.assertTrue(any("Decision impact" in warning for warning in warnings))
        self.assertTrue(any("evidence-specific" in warning for warning in warnings))
        self.assertTrue(any("FP" in warning and "exactly one cluster" in warning for warning in warnings))

    def test_ucp_use_case_rows_require_count_provenance(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "08_UCP"
        sheet.append(
            [
                "分類",
                "項目",
                "複雑度",
                "数量",
                "重み",
                "UCP",
                "Source status",
                "Source locator",
                "根拠",
                "メモ",
            ]
        )
        sheet.append(["Use case", "Related flows", "Average", 6, 10, 60, "", "", "inferred", ""])

        warnings = formatter.check_functional_count_provenance(workbook)

        self.assertTrue(any("Source status" in warning for warning in warnings))
        self.assertTrue(any("Source locator" in warning for warning in warnings))

    @staticmethod
    def workbook_with_ucp_reconciliation(derived, untraced, ratio, status):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "08_UCP"
        sheet.append(["分類", "項目", "複雑度", "数量", "重み", "UCP", "Source status", "Source locator", "根拠", "メモ"])
        sheet.append(["Use case", "Published use cases", "Simple", 9, 5, 45, "explicit", "requirements:use-cases", "source", ""])
        sheet.append([])
        sheet.append(["Metric", "Explicit count", "Derived count", "Untraced inferred", "Inflation ratio", "Guard status"])
        sheet.append(["UUCP", 57, derived, untraced, ratio, status])
        return workbook

    def test_ucp_numeric_count_inflation_is_strict_warning(self):
        workbook = self.workbook_with_ucp_reconciliation(
            94, 37, (94 - 57) / 57, "STOP_UNTRACED_COUNT"
        )

        warnings = formatter.check_functional_count_provenance(workbook)

        self.assertTrue(any("STOP_UNTRACED_COUNT" in warning and "64.9%" in warning for warning in warnings))

    def test_ucp_matching_explicit_count_passes_provenance_qa(self):
        workbook = self.workbook_with_ucp_reconciliation(57, 0, 0, "PASS")

        self.assertEqual([], formatter.check_functional_count_provenance(workbook))


if __name__ == "__main__":
    unittest.main()
